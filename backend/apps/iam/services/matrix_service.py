import logging
from io import BytesIO
from typing import Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import quote_sheetname
from django.db import transaction
from django.core.exceptions import ValidationError
from collections import Counter

from apps.iam.models.rbac_models import Permission, Role, RolePermission, RbacSubModule, RbacModule

logger = logging.getLogger(__name__)

class MatrixService:
    """
    Service for generating and parsing the Role-Permission Matrix Excel template.
    Now supports Module-wise Sheets and Auto-Detects Sub-Modules based on selected permissions.
    """

    @classmethod
    def export_matrix(cls) -> BytesIO:
        """
        Generate an Excel workbook containing all active permissions, grouped into separate sheets per module.
        Returns a BytesIO object containing the Excel file.
        """
        wb = Workbook()
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Define Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Indigo 600
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Fetch Taxonomy
        modules = RbacModule.objects.all().order_by('code')
        sub_modules = {sm.code: sm for sm in RbacSubModule.objects.all()}
        permissions = list(Permission.objects.filter(is_active=True).order_by('module', 'name'))

        for module in modules:
            # Permissions for this module's sub-modules
            module_sub_codes = [sm.code for sm in sub_modules.values() if sm.parent_module_id == module.id]
            mod_permissions = [p for p in permissions if p.module in module_sub_codes or p.module == module.code]
            
            if not mod_permissions:
                continue

            # Create Sheet, max length for sheet title is 31
            sheet_title = (module.title or module.code)[:31]
            ws = wb.create_sheet(title=sheet_title)

            # Setup Headers
            headers = [
                "Module Name",
                "Sub-Module Name",
                "Permission Title",
                "Permission Code (DO NOT EDIT)"
            ]
            
            num_example_roles = 5
            role_cols_start = len(headers) + 1

            # Row 1: Role Names
            ws.append([""] * len(headers) + [f"Role {i+1}" for i in range(num_example_roles)])
            # Row 2: Department Names
            ws.append([""] * len(headers) + [f"Dept {i+1}" for i in range(num_example_roles)])
            # Row 3: Main Headers
            ws.append(headers + ["Type 'X' below to assign" for _ in range(num_example_roles)])

            # Apply styles
            for col_idx in range(1, role_cols_start + num_example_roles):
                # Row 1: Role
                cell1 = ws.cell(row=1, column=col_idx)
                cell1.font = Font(bold=True, color="111827")
                cell1.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
                cell1.border = thin_border
                cell1.alignment = align_center

                # Row 2: Dept
                cell2 = ws.cell(row=2, column=col_idx)
                cell2.font = Font(italic=True, color="6B7280")
                cell2.fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
                cell2.border = thin_border
                cell2.alignment = align_center

                # Row 3: Column Headers
                cell3 = ws.cell(row=3, column=col_idx)
                cell3.font = header_font
                cell3.fill = header_fill
                cell3.border = thin_border
                cell3.alignment = align_center

            # Labels for Row 1, 2
            ws.cell(row=1, column=role_cols_start - 1).value = "Role Name ->"
            ws.cell(row=1, column=role_cols_start - 1).font = Font(bold=True)
            ws.cell(row=1, column=role_cols_start - 1).alignment = Alignment(horizontal="right")
            ws.cell(row=2, column=role_cols_start - 1).value = "Department (Optional) ->"
            ws.cell(row=2, column=role_cols_start - 1).font = Font(italic=True)
            ws.cell(row=2, column=role_cols_start - 1).alignment = Alignment(horizontal="right")

            # Adjust Column Widths
            ws.column_dimensions['A'].width = 20 # Module
            ws.column_dimensions['B'].width = 25 # Sub-Module
            ws.column_dimensions['C'].width = 40 # Title
            ws.column_dimensions['D'].width = 30 # Code
            for i in range(num_example_roles):
                ws.column_dimensions[chr(ord('E') + i)].width = 20

            current_row = 4
            for perm in mod_permissions:
                sub_module_name = perm.module
                if perm.module in sub_modules:
                    sm = sub_modules[perm.module]
                    sub_module_name = sm.title
                
                row_data = [
                    module.title,
                    sub_module_name,
                    perm.name,
                    perm.code
                ]
                
                ws.append(row_data)
                
                # Apply styles to data rows
                for col_idx in range(1, len(row_data) + 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = align_left
                
                for col_idx in range(role_cols_start, role_cols_start + num_example_roles):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = align_center

                current_row += 1

            # Freeze Panes
            ws.freeze_panes = 'E4' # Freeze headers and first 4 columns

        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @classmethod
    @transaction.atomic
    def import_matrix(cls, business_unit_id: str, file_obj, user_id: str):
        """
        Parse the uploaded Excel file and create/update Roles and Permissions.
        Iterates over all sheets.
        Row 1: Role Names
        Row 2: Department Names
        Row 3: Headers
        Row 4+: Permissions
        """
        wb = load_workbook(filename=file_obj, data_only=True)
        
        created_count = 0
        updated_count = 0

        valid_permission_codes = set(Permission.objects.filter(is_active=True).values_list('code', flat=True))
        perm_map = {p.code: p for p in Permission.objects.filter(code__in=valid_permission_codes)}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 1. Parse Roles from Row 1, 2
            role_cols_start = 5 # Column E
            max_col = ws.max_column
            
            roles_to_process = [] 

            for col in range(role_cols_start, max_col + 1):
                role_name_cell = ws.cell(row=1, column=col).value
                dept_name_cell = ws.cell(row=2, column=col).value

                if role_name_cell and str(role_name_cell).strip() and not str(role_name_cell).startswith('Role '):
                    roles_to_process.append({
                        'col': col,
                        'name': str(role_name_cell).strip(),
                        'dept': str(dept_name_cell).strip() if dept_name_cell and not str(dept_name_cell).startswith('Dept ') else None,
                        'permission_codes': []
                    })

            if not roles_to_process:
                continue

            # 2. Parse Permissions
            max_row = ws.max_row
            
            for row in range(4, max_row + 1):
                perm_code_cell = ws.cell(row=row, column=4).value
                if not perm_code_cell:
                    continue
                    
                perm_code = str(perm_code_cell).strip()
                if perm_code not in valid_permission_codes:
                    continue 

                for role_data in roles_to_process:
                    cell_value = ws.cell(row=row, column=role_data['col']).value
                    if cell_value:
                        val = str(cell_value).strip().lower()
                        if val in ['x', 'yes', 'y', '1', 'true', 'v']:
                            role_data['permission_codes'].append(perm_code)

            # 3. Create or Update Roles in DB
            for role_data in roles_to_process:
                # Auto-detect Sub-Module based on majority of assigned permissions
                detected_sub_module = None
                if role_data['permission_codes']:
                    assigned_perms = [perm_map[code] for code in role_data['permission_codes'] if code in perm_map]
                    if assigned_perms:
                        # Count the occurrences of each sub-module
                        sub_module_counts = Counter([p.module for p in assigned_perms if p.module])
                        if sub_module_counts:
                            # Get the most common sub-module
                            detected_sub_module = sub_module_counts.most_common(1)[0][0]

                role, created = Role.objects.get_or_create(
                    business_unit_id=business_unit_id,
                    name=role_data['name'],
                    defaults={
                        'department_name': role_data['dept'],
                        'module_code': detected_sub_module,
                        'role_type': Role.RoleType.CUSTOM,
                        'is_default': False,
                        'created_by_id': user_id,
                        'updated_by_id': user_id
                    }
                )
                
                if not created:
                    needs_save = False
                    if role_data['dept'] and role.department_name != role_data['dept']:
                        role.department_name = role_data['dept']
                        needs_save = True
                    if detected_sub_module and role.module_code != detected_sub_module:
                        role.module_code = detected_sub_module
                        needs_save = True
                        
                    if needs_save:
                        role.updated_by_id = user_id
                        role.save(update_fields=['department_name', 'module_code', 'updated_by_id'])
                        
                    updated_count += 1
                else:
                    created_count += 1

                # Sync Permissions
                perms_to_add = [perm_map[code] for code in role_data['permission_codes'] if code in perm_map]
                RolePermission.objects.filter(role=role).delete()
                
                rp_objects = [
                    RolePermission(
                        role=role,
                        permission=p,
                        created_by_id=user_id,
                        updated_by_id=user_id
                    ) for p in perms_to_add
                ]
                RolePermission.objects.bulk_create(rp_objects)

        if created_count == 0 and updated_count == 0:
            raise ValidationError("No valid roles found in any sheet. Please provide role names in the columns.")

        return {
            "success": True,
            "roles_created": created_count,
            "roles_updated": updated_count
        }
