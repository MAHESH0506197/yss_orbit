import pytest
import uuid
import io
import openpyxl
from django.core.files.uploadedfile import SimpleUploadedFile
from apps.hrms.models import Employee, Department, Designation
from apps.hrms.models.employee_import_session import EmployeeImportSession
from apps.hrms.services.employee_import_service import EmployeeImportService

@pytest.mark.django_db
class TestEmployeeImportService:
    def test_validation_duplicate_code(self, tenant_bu, default_user):
        Employee.objects.create(
            business_unit_id=tenant_bu.id,
            employee_code="EMP-123",
            first_name="John",
            last_name="Doe",
            work_email="john@example.com",
            date_of_joining="2024-01-01",
        )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Employee Code", "First Name", "Last Name", "Official Email", "Department", "Designation", "Employment Type", "Employment Status", "Joining Date (YYYY-MM-DD)"])
        ws.append(["EMP-123", "Jane", "Smith", "jane@example.com", "IT", "Dev", "FULL_TIME", "ACTIVE", "2024-01-01"])
        
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        
        file = SimpleUploadedFile("test.xlsx", out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        session = EmployeeImportSession.objects.create(
            business_unit_id=tenant_bu.id,
            uploaded_by=default_user,
            file_path=file,
            original_file_name="test.xlsx"
        )
        
        EmployeeImportService.validate_session(session)
        assert session.status == EmployeeImportSession.Status.FAILED
        assert session.error_rows == 1
        assert session.error_grid[0]["error"] == "Already exists in database"

    def test_validation_invalid_status(self, tenant_bu, default_user):
        # Create Dept/Desig so validation passes them and reaches the status check
        dept = Department.objects.create(business_unit_id=tenant_bu.id, name="IT")
        Designation.objects.create(business_unit_id=tenant_bu.id, name="Dev", department=dept)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Employee Code", "First Name", "Last Name", "Official Email", "Department", "Designation", "Employment Type", "Employment Status", "Joining Date (YYYY-MM-DD)"])
        ws.append(["EMP-999", "Jane", "Smith", "jane@example.com", "IT", "Dev", "FULL_TIME", "INVALID_STATUS", "2024-01-01"])
        
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        
        file = SimpleUploadedFile("test.xlsx", out.read())
        session = EmployeeImportSession.objects.create(
            business_unit_id=tenant_bu.id,
            uploaded_by=default_user,
            file_path=file,
            original_file_name="test.xlsx"
        )
        
        EmployeeImportService.validate_session(session)
        assert session.status == EmployeeImportSession.Status.FAILED
        assert session.error_grid[0]["error"] == "Invalid status"


    def test_execute_session_success(self, tenant_bu, default_user):
        # Create dependencies
        dept = Department.objects.create(business_unit_id=tenant_bu.id, name="IT")
        desig = Designation.objects.create(business_unit_id=tenant_bu.id, name="Dev", department=dept)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Employee Code", "First Name", "Last Name", "Official Email", "Department", "Designation", "Employment Type", "Employment Status", "Joining Date (YYYY-MM-DD)"])
        for i in range(1, 101):
            ws.append([f"EMP-{i}", "Jane", f"Smith{i}", f"jane{i}@example.com", "IT", "Dev", "FULL_TIME", "ACTIVE", "2024-01-01"])
            
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        
        file = SimpleUploadedFile("test.xlsx", out.read())
        session = EmployeeImportSession.objects.create(
            business_unit_id=tenant_bu.id,
            uploaded_by=default_user,
            file_path=file,
            original_file_name="test.xlsx"
        )
        
        EmployeeImportService.validate_session(session)
        assert session.status == EmployeeImportSession.Status.VALIDATED
        assert session.valid_rows == 100
        
        EmployeeImportService.execute_session(session)
        assert session.status == EmployeeImportSession.Status.COMPLETED
        assert Employee.objects.filter(business_unit_id=tenant_bu.id).count() == 100
