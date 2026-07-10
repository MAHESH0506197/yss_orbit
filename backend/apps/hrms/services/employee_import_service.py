from typing import Any
import io
import uuid
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment
from django.db import transaction
from apps.hrms.models import Department, Designation, Employee
from apps.hrms.models.employee_import_session import EmployeeImportSession

class EmployeeImportService:
    @staticmethod
    def generate_template(bu_id: uuid.UUID) -> bytes:
        """
        Generates a 3-sheet Excel template for employee bulk import.
        Sheet 1: Employees (Data Entry)
        Sheet 2: Instructions
        Sheet 3: Reference Data (Departments, Designations)
        """
        wb = openpyxl.Workbook()
        
        # Sheet 1: Employees
        ws_emp = wb.active
        ws_emp.title = "Employees"
        
        columns = [
            "Employee Code", "First Name", "Last Name", "Official Email", 
            "Mobile", "Department", "Designation", "Employment Type", 
            "Employment Status", "Manager Code", "Joining Date (YYYY-MM-DD)",
            "PAN", "Aadhaar", "IFSC", "Basic Salary", "CTC"
        ]
        
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header_text in enumerate(columns, 1):
            cell = ws_emp.cell(row=1, column=col_num, value=header_text)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            ws_emp.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
        
        # Sheet 2: Instructions
        ws_inst = wb.create_sheet("Instructions")
        instructions = [
            ["Column", "Required", "Format", "Rules"],
            ["Employee Code", "Yes", "Text", "Must be unique across the organization."],
            ["First Name", "Yes", "Text", ""],
            ["Last Name", "Yes", "Text", ""],
            ["Official Email", "Yes", "Email", "Must be valid format and unique."],
            ["Mobile", "Yes", "Number", "10 digit number."],
            ["Department", "Yes", "Text", "Must exactly match a department name from Reference Data."],
            ["Designation", "Yes", "Text", "Must exactly match a designation name from Reference Data."],
            ["Employment Type", "Yes", "Text", "FULL_TIME, PART_TIME, CONTRACT, INTERN"],
            ["Employment Status", "Yes", "Text", "ACTIVE, ON_LEAVE, NOTICE_PERIOD, TERMINATED, RESIGNED"],
            ["Manager Code", "No", "Text", "Employee Code of the reporting manager."],
            ["Joining Date", "Yes", "YYYY-MM-DD", "Must not be in the future."],
            ["PAN", "No", "Text", "ABCDE1234F format."],
            ["Aadhaar", "No", "Text", "12 digits."],
        ]
        
        for row_num, row_data in enumerate(instructions, 1):
            for col_num, cell_value in enumerate(row_data, 1):
                cell = ws_inst.cell(row=row_num, column=col_num, value=cell_value)
                if row_num == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                ws_inst.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25
        
        # Sheet 3: Reference Data
        ws_ref = wb.create_sheet("Reference Data")
        ws_ref.cell(row=1, column=1, value="Departments").font = Font(bold=True)
        ws_ref.cell(row=1, column=2, value="Designations").font = Font(bold=True)
        
        depts = list(Department.objects.filter(business_unit_id=bu_id).values_list("name", flat=True))
        desigs = list(Designation.objects.filter(business_unit_id=bu_id).values_list("name", flat=True))
        
        for i, dept in enumerate(depts, 2):
            ws_ref.cell(row=i, column=1, value=dept)
            
        for i, desig in enumerate(desigs, 2):
            ws_ref.cell(row=i, column=2, value=desig)
            
        ws_ref.column_dimensions["A"].width = 30
        ws_ref.column_dimensions["B"].width = 30
        
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.read()

    @staticmethod
    def validate_session(session: EmployeeImportSession) -> None:
        """
        Parses the uploaded Excel file and validates every row against
        DDD business rules, uniqueness constraints, and relationships.
        Updates the session with the results.
        """
        session.status = EmployeeImportSession.Status.VALIDATING
        session.save(update_fields=["status"])

        import re
        from datetime import datetime
        from django.utils import timezone

        bu_id = session.business_unit_id
        
        # Pre-fetch lookup data
        depts = {d.name.lower(): d.id for d in Department.objects.filter(business_unit_id=bu_id)}
        desigs = {d.name.lower(): d.id for d in Designation.objects.filter(business_unit_id=bu_id)}
        existing_emp_codes = set(Employee.objects.filter(business_unit_id=bu_id).values_list("employee_code", flat=True))
        existing_emails = set(Employee.objects.filter(business_unit_id=bu_id).values_list("work_email", flat=True))
        
        wb = openpyxl.load_wsgi(session.file_path) if hasattr(openpyxl, "load_wsgi") else openpyxl.load_workbook(session.file_path.path, data_only=True)
        ws = wb.active
        
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value else "")
            
        required_headers = ["Employee Code", "First Name", "Last Name", "Official Email", "Department", "Designation", "Employment Type", "Employment Status", "Joining Date (YYYY-MM-DD)"]
        
        # File-Level Validation
        missing_headers = [h for h in required_headers if h not in headers]
        if missing_headers:
            session.status = EmployeeImportSession.Status.FAILED
            session.error_grid = [{"row": 1, "column": "Headers", "value": "", "expected": ", ".join(missing_headers), "error": "Missing required columns"}]
            session.save(update_fields=["status", "error_grid"])
            return

        col_map = {h: i for i, h in enumerate(headers)}
        
        error_grid = []
        valid_rows_data = []
        
        file_codes = set()
        file_emails = set()
        
        today = timezone.now().date()
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            # Check if row is completely empty
            if all(v is None for v in row):
                continue
                
            def get_val(col_name):
                idx = col_map.get(col_name)
                return str(row[idx]).strip() if idx is not None and row[idx] is not None else ""
                
            code = get_val("Employee Code")
            fname = get_val("First Name")
            lname = get_val("Last Name")
            email = get_val("Official Email")
            dept_name = get_val("Department")
            desig_name = get_val("Designation")
            emp_type = get_val("Employment Type")
            emp_status = get_val("Employment Status")
            joining_date = get_val("Joining Date (YYYY-MM-DD)")
            mobile = get_val("Mobile")
            
            row_errors = []
            
            # 1. Required Fields
            if not code: row_errors.append({"column": "Employee Code", "value": "", "expected": "Text", "error": "Required field"})
            if not fname: row_errors.append({"column": "First Name", "value": "", "expected": "Text", "error": "Required field"})
            if not email: row_errors.append({"column": "Official Email", "value": "", "expected": "Email", "error": "Required field"})
            if not dept_name: row_errors.append({"column": "Department", "value": "", "expected": "Text", "error": "Required field"})
            
            # 2. Duplicates
            if code in existing_emp_codes:
                row_errors.append({"column": "Employee Code", "value": code, "expected": "Unique", "error": "Already exists in database"})
            elif code in file_codes:
                row_errors.append({"column": "Employee Code", "value": code, "expected": "Unique", "error": "Duplicate in file"})
            if code: file_codes.add(code)
                
            if email in existing_emails:
                row_errors.append({"column": "Official Email", "value": email, "expected": "Unique", "error": "Already exists in database"})
            elif email in file_emails:
                row_errors.append({"column": "Official Email", "value": email, "expected": "Unique", "error": "Duplicate in file"})
            if email: file_emails.add(email)
            
            # 3. Relationships
            dept_id = depts.get(dept_name.lower())
            if dept_name and not dept_id:
                row_errors.append({"column": "Department", "value": dept_name, "expected": "Existing Department", "error": "Department not found in this Business Unit"})
                
            desig_id = desigs.get(desig_name.lower())
            if desig_name and not desig_id:
                row_errors.append({"column": "Designation", "value": desig_name, "expected": "Existing Designation", "error": "Designation not found in this Business Unit"})
                
            # 4. Enums
            if emp_status and emp_status not in [c[0] for c in Employee.EmploymentStatus.choices]:
                row_errors.append({"column": "Employment Status", "value": emp_status, "expected": "ACTIVE, ON_LEAVE, etc", "error": "Invalid status"})
                
            # 5. Cross-field Date Validation
            try:
                if joining_date and " " in joining_date: joining_date = joining_date.split()[0]
                j_date = datetime.strptime(joining_date, "%Y-%m-%d").date() if joining_date else None
                if j_date and j_date > today:
                    row_errors.append({"column": "Joining Date", "value": joining_date, "expected": "<= Today", "error": "Joining date cannot be in the future"})
            except ValueError:
                row_errors.append({"column": "Joining Date", "value": joining_date, "expected": "YYYY-MM-DD", "error": "Invalid date format"})
                j_date = None
            
            if row_errors:
                for err in row_errors:
                    err["row"] = row_num
                error_grid.extend(row_errors)
            else:
                valid_rows_data.append({
                    "employee_code": code,
                    "first_name": fname,
                    "last_name": lname,
                    "work_email": email,
                    "phone": mobile,
                    "department_id": str(dept_id) if dept_id else None,
                    "designation_id": str(desig_id) if desig_id else None,
                    "employment_type": emp_type,
                    "employment_status": emp_status,
                    "date_of_joining": str(j_date) if j_date else None
                })
                
        total_rows = len(valid_rows_data) + len(set(e["row"] for e in error_grid))
        
        session.total_rows = total_rows
        session.valid_rows = len(valid_rows_data)
        session.error_rows = len(set(e["row"] for e in error_grid))
        
        session.validation_summary = {
            "valid_data": valid_rows_data if not error_grid else []
        }
        session.error_grid = error_grid
        
        session.status = EmployeeImportSession.Status.FAILED if error_grid else EmployeeImportSession.Status.VALIDATED
        session.save()

    @staticmethod
    @transaction.atomic
    def execute_session(session: EmployeeImportSession) -> None:
        if session.status != EmployeeImportSession.Status.VALIDATED:
            raise ValueError("Session is not validated or has errors.")
            
        session.status = EmployeeImportSession.Status.IMPORTING
        session.save(update_fields=["status"])
        
        valid_data = session.validation_summary.get("valid_data", [])
        
        employees_to_create = []
        for row in valid_data:
            employees_to_create.append(Employee(
                business_unit_id=session.business_unit_id,
                created_by_id=session.uploaded_by_id,
                **row
            ))
            
        Employee.objects.bulk_create(employees_to_create, batch_size=500)
        
        session.status = EmployeeImportSession.Status.COMPLETED
        session.save(update_fields=["status"])

    @staticmethod
    def generate_error_report(session: EmployeeImportSession) -> bytes:
        if not session.error_grid:
            return b""
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Errors"
        
        headers = ["Row", "Column", "Invalid Value", "Expected", "Error Message"]
        header_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header_text)
            cell.fill = header_fill
            cell.font = header_font
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 25
            
        for row_num, err in enumerate(session.error_grid, 2):
            ws.cell(row=row_num, column=1, value=err.get("row", ""))
            ws.cell(row=row_num, column=2, value=err.get("column", ""))
            ws.cell(row=row_num, column=3, value=err.get("value", ""))
            ws.cell(row=row_num, column=4, value=err.get("expected", ""))
            ws.cell(row=row_num, column=5, value=err.get("error", ""))
            
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        return out.read()
